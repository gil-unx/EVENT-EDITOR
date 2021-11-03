import json
import openpyxl
from PIL import Image
from TEXTEDITOR import *
def repackCode(s):
    comand = s[1]
    try:
        code = inv_cod[comand]
    except:
        code = int(comand, 16)
    b = 0
    v0 = int(s[2])
    v1 = int(s[3])
    v2 = int(s[4])
    if comand == "MOVE-NPC":
        a0 = int(s[5][7:-1])
        a1 = int(inv_direction[s[6][1:-1]])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][10:-1])
    elif comand == "OPENTEXTBOX":
        a0 = int(s[5][8:-1])
        a1 = int(s[6][6:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "SETNPCPOS":
        a0 = int(s[5][7:-1])
        a1 = int(s[6][2:-1])
        a2 = int(s[7][2:-1])
        a3 = int(s[8][2:-1])
    elif comand == "SETNPCDIR":
        a0 = int(s[5][7:-1])
        a1 = int(s[6][10:-1])
        a2 = int(inv_direction[s[7][1:-1]])
        a3 = int(s[8][1:-1])
    elif comand == "CHANGEMAP":
        a0 = inv_mapdict[s[5][1:-1]]
        a1 = int(s[6][1:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "MOVE-CAM":
        a0 = int(s[5])
        a1 = int(s[6])
        a2 = int(s[7])
        a3 = int(s[8])
    elif comand == "SOUND-MV":
        a0 = int(s[5][9:-1])
        a1 = int(s[6][1:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "BGM":
        a0 = int(s[5][4:-1])
        a1 = int(s[6][1:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "LIGHT":
        a0 = int(s[5])
        a1 = int(s[6])
        a2 = int(s[7])
        a3 = int(s[8])
    elif comand == "OPEN-DOOR":
        a0 = int(s[5][5:-1])
        a1 = int(s[6][1:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "MOVE-CAM-TO-NPC":
        a0 = int(s[5][4:-1])
        a1 = int(s[6][1:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "SET-OBJ-POS":
        a0 = int(s[5][4:-1])
        a1 = int(s[6][2:-1])
        a2 = int(s[7][2:-1])
        a3 = int(s[8][2:-1])

    elif comand == "SETMCPOS":
        a0 = int(s[5][5:-1])
        a1 = int(s[6][2:-1])
        a2 = int(s[7][2:-1])
        a3 = int(s[8][2:-1])

    elif comand == "MOVE-MC":
        a0 = int(s[5][5:-1])
        a1 = int(inv_direction[s[6][1:-1]])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][4:-1])

    elif comand == "MC-ACTION":
        a0 = int(s[5])
        a1 = int(s[6])
        a2 = int(s[7])
        a3 = int(s[8])
    elif comand == "SET-MC-DIR":
        a0 = int(inv_direction[s[5][1:-1]])
        a1 = int(s[6][1:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "ADVANCE-TIME":
        a0 = int(s[5][5:-1])
        a1 = int(s[6][7:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "MOVE-NPC(SP1)":
        a0 = int(s[5][7:-1])
        a1 = int(inv_direction[s[6][1:-1]])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][10:-1])
        # input("ddddd")
    # input()

    else:
        a0 = int(s[5])
        a1 = int(s[6])
        a2 = int(s[7])
        a3 = int(s[8])
    return struct.pack("<4bHh2i", code, v0, v1, v2, a0, a1, a2, a3)


def Encode(jso, npcName):
    f = Br(io.BytesIO())
    f.writeUint32(0xE)
    infoSize = [8, 4, 12, 4, 10]
    newOff = []
    off = 0
    for node, inf in zip(jso[npcName], infoSize):
        newOff.append(off + 0x24)
        lin = len(jso[npcName][node])
        off += (lin * inf)
        if node == "ANIMATION":
            lin = 0
            for j in jso[npcName]["ANIMATION"]:
                lin += len(j)
        f.writeUint16(lin)
    f.writeUint16(0)
    for o in newOff:
        f.writeUint32(o)
    for node in jso[npcName]["RAM_INFO"]:
        f.write(struct.pack("<hhhh", *node))
    for node in jso[npcName]["DICT"]:
        f.write(struct.pack("<hh", *node))
    for node in jso[npcName]["ANIM_INFO"]:
        f.write(struct.pack("<hhhhbbbb", *node))
    for node in jso[npcName]["POS"]:
        f.write(struct.pack("<hh", *node))
    for node in jso[npcName]["ANIMATION"]:
        for cNode in node:
            f.write(struct.pack("<hhhhh", *cNode))
    return f.getdata()


class Event:
    def __init__(self, data):
        self.data = Br(io.BytesIO(data))

    def extract(self):
        f = self.data
        f.seek(0xd74)
        pallets = []
        tiles = []
        for i in range(4):
            f.read(12)
            pallets.append(f.read(0x200))
        for i in range(4):
            tileSize = f.readUint16()
            f.read(10)
            tiles.append(f.read(tileSize))
        for i in range(4):
            bpal = Pal(pallets[i])
            palet = bpal.getpal()
            texture = Tile(tiles[i], 64, 240)
            datPng = texture.getpng(palet)
            fileSave(datPng, "HHMBTNI/EVENT_BASE/NPC_SPRITE{0}.png".format(i + 1))

    def insert(self, slotNpc, sheet,vXpos,vYpos,palXpos,palYpos):
        # <header>
        newEv = Br(io.BytesIO())
        newEv.write(b"\x80\x81")
        newEv.writeUint16(len(slotNpc) + 2)
        filesOffset = []
        files = Br(io.BytesIO())
        for i in range(len(slotNpc)):
            newEv.writeUint8(i + 1)
            newEv.writeUint8(32)
        newEv.writeUint16(0x8000)
        newEv.writeUint16(0x4000)
        newEv.wpad(0x34)
        # <init>
        tiles = Br(io.BytesIO())
        indexNpc = 0
        totalSpr = 0
        for jsName in slotNpc:
            npcName = jsName[12:]

            js = json.loads(open(jsName + "/INFO.json", "r").read())
            for i in range(len(js[npcName]["RAM_INFO"])):
                x, y, w, h = js[npcName]["RAM_INFO"][i]
                if js[npcName]["RAM_INFO"][i][3] != 1:
                    spr = Image.open(jsName + "/SPRITE/{0:04d}.png".format(i))
                    if (vYpos + h) > 500:
                        vXpos += js[npcName]["MAX_WIDTH"]
                        vYpos = 256
                    js[npcName]["RAM_INFO"][i] = vXpos, vYpos , w, h
                    tile = Br(open(jsName + "/BINARY/{0:04d}.bin".format(i), "r+b"))
                    tile.read(4)
                    tile.write(
                        struct.pack("<hhhh", vXpos , vYpos , w, h))
                    tiles.write(tile.getdata())
                    tiles.wpad(4)
                    vYpos += h
                else:
                    js[npcName]["RAM_INFO"][i] = palXpos, palYpos, w, h
                    tile = Br(open(jsName + "/BINARY/{0:04d}.bin".format(i), "r+b"))
                    tile.read(2)
                    tile.write(b"\x00\x04")
                    tile.write(struct.pack("<hhhh", palXpos, palYpos, w, h))
                    tiles.write(tile.getdata())
                    tiles.wpad(4)
                    palYpos-=1
                totalSpr += 1
            filesOffset.append(files.tell())
            files.write(Encode(js, npcName))
            indexNpc += js[npcName]["SLOT"]
            vXpos+=js[npcName]["MAX_WIDTH"]
            vYpos = 256

        filesOffset.append(files.tell())
        files.write(struct.pack("<HH", 0x100d, totalSpr) + tiles.getdata())
        filesOffset.append(files.tell())

        i = 1
        while True:
            num = sheet.cell(row=i, column=1).value
            s = []
            for j in range(9):
                val = sheet.cell(row=i, column=j + 1).value
                s.append(val)

            try:
                if s[1] == "END":
                    files.write(b"\xff\x00\x00\x00\x00")
                    break
                else:
                	newCode = repackCode(s)
                	files.write(newCode)
                

            except:
                print(s[1])
                print("error di line :{0}".format(num))
                sys.exit()

            i += 1

        filesOffset.append(files.tell())
        for i in filesOffset:
            newEv.writeUint32(i)
        newEv.write(files.getdata())

        return newEv.getdata()

def Main(npc,vX,vY,pX,pY):
    filename = "HHMBTNI/EVENT_BASE/COMMAND.xlsx"
    npcs = []
    wb = openpyxl.load_workbook(filename, data_only=True)
    sheet = wb.active
    i = 1
    while i < npc:
        npcname = sheet.cell(row=1 + i, column=12).value
        if npcname:
            npcs.append("HHMBTNI/NPC/" + npcname)
        i += 1

    data = open("HHMBTNI/EVENT_BASE.bin", "rb").read()
    print("REPACK EVENT...")
    ev = Event(data)
    buffEvent = ev.insert(npcs, sheet, vX,vY,pX,pY)
    newEvent = io.BytesIO(buffEvent)
    eventSize = len(buffEvent)
    if eventSize > 0x32000:
        print("Max size terlewati!!\nInsert Gagal")
        sys.exit()

    print("REPACK EVENT SIP!")
    iso = open("isos/HHMBTNI_BASE.iso", "r+b")
    iso.seek(0x85e1f98)
    while chunk := newEvent.read(0x800):
        iso.write(chunk)
        iso.read(0x130)
    ins(iso)


# v1 -64,256,-64,511
# Standarddd 320,256,384,159


if __name__ == "__main__":
    Main( 4, 320,256,384,159)
    #Main(20,-64,256,-64,511)

