import traceback
import gzip
import json
import glob
import openpyxl
from FUN import *
from PIL import Image
from PyTexturePacker import main
from TEXTEDITOR import *
dicNpc = {}
dicItem = {}
def repackCode(s):
    comand = s[1]
    try:
        code = inv_cod[comand]
    except:
        code = int(comand, 16)
    b = 0
    v0 = int(s[2])
    v1 = int(s[3])
    v2 = int(s[4])
    if comand == "MOVE-NPC":
        a0 = int(dicNpc[s[5][1:-1]])
        a1 = int(inv_direction[s[6][1:-1]])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][10:-1])
    elif comand == "OPENTEXTBOX":
        a0 = int(s[5][8:-1])
        a1 = int(s[6][6:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "SETNPCPOS":
        a0 = int(dicNpc[s[5][1:-1]])
        a1 = int(s[6][2:-1])
        a2 = int(s[7][2:-1])
        a3 = int(s[8][2:-1])
    elif comand == "SETNPCDIR":
        a0 = int(dicNpc[s[5][1:-1]])
        a1 = int(s[6][10:-1])
        a2 = int(inv_direction[s[7][1:-1]])
        a3 = int(s[8][1:-1])
    elif comand == "CHANGEMAP":
        a0 = inv_mapdict[s[5][1:-1]]
        a1 = int(s[6][1:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "MOVE-CAM-TO-POS":
        a0 = int(s[5][1:-1])
        a1 = int(s[6][2:-1])
        a2 = int(s[7][2:-1])
        a3 = int(s[8][2:-1])
    elif comand == "SOUND-MV":
        a0 = int(s[5][9:-1])
        a1 = int(s[6][1:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "BGM":
        a0 = int(s[5][4:-1])
        a1 = int(s[6][1:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "LIGHT":
        a0 = int(s[5])
        a1 = int(s[6])
        a2 = int(s[7])
        a3 = int(s[8])
    elif comand == "MC-HANDLING-ITEM":
        a0 = DICT_ITEM[s[5][1:-1]]
        a1 = int(s[6][1:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "OPEN-DOOR":
        a0 = int(s[5][5:-1])
        a1 = int(s[6][1:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "FLOATING-ICON":
        a0 = int(dicNpc[s[5][1:-1]])
        a1 = icon_dic[s[6][1:-1]]
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "MOVE-CAM-TO-NPC":
        a0 = int(dicNpc[s[5][1:-1]])
        a1 = int(s[6][1:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "MOVE-CAM-TO-MC":
        a0 = -2#int(s[5][5:-1])
        a1 = int(s[6][1:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "SET-ITEM-POS":
        a0 = dicItem["{0:04d}".format(DICT_ITEM[s[5][1:-1]])]
        a1 = int(s[6][2:-1])
        a2 = int(s[7][2:-1])
        a3 = int(s[8][2:-1])

    elif comand == "SETMCPOS":
        a0 = int(s[5][5:-1])
        a1 = int(s[6][2:-1])
        a2 = int(s[7][2:-1])
        a3 = int(s[8][2:-1])

    elif comand == "MOVE-MC":
        a0 = int(s[5][5:-1])
        a1 = int(inv_direction[s[6][1:-1]])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][4:-1])

    elif comand == "MC-ACTION":
        a0 = int(s[5][1:-1])
        a1 = mcActDict[s[6]]
        a2 = int(inv_direction[s[7][1:-1]])
        a3 = int(s[8][1:-1])
    elif comand == "SET-MC-DIR":
        a0 = int(inv_direction[s[5][1:-1]])
        a1 = int(s[6][1:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "DELAY":
        a0 = int(s[5][1:-1])
        a1 = int(s[6][1:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "ADVANCE-TIME":
        a0 = int(s[5][5:-1])
        a1 = int(s[6][7:-1])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][1:-1])
    elif comand == "MOVE-NPC(SP1)":
        a0 = int(s[5][7:-1])
        a1 = int(inv_direction[s[6][1:-1]])
        a2 = int(s[7][1:-1])
        a3 = int(s[8][10:-1])
    else:
        a0 = int(s[5])
        a1 = int(s[6])
        a2 = int(s[7])
        a3 = int(s[8])
    return struct.pack("<4bhh2i", code, v0, v1, v2, a0, a1, a2, a3)


def Encode(jso, npcName):
    f = Br(io.BytesIO())
    f.writeUint32(0xE)
    infoSize = [8, 4, 12, 4, 10]
    newOff = []
    off = 0
    for node, inf in zip(jso[npcName], infoSize):
        newOff.append(off + 0x24)
        lin = len(jso[npcName][node])
        off += (lin * inf)
        if node == "ANIMATION":
            lin = 0
            for j in jso[npcName]["ANIMATION"]:
                lin += len(j)
        f.writeUint16(lin)
    f.writeUint16(0)
    for o in newOff:
        f.writeUint32(o)
    for node in jso[npcName]["RAM_INFO"]:
        f.write(struct.pack("<hhhh", *node))
    for node in jso[npcName]["DICT"]:
        f.write(struct.pack("<hh", *node))
    for node in jso[npcName]["ANIM_INFO"]:
        f.write(struct.pack("<hhhhbbbb", *node))
    for node in jso[npcName]["POS"]:
        f.write(struct.pack("<hh", *node))
    for node in jso[npcName]["ANIMATION"]:
        for cNode in node:
            f.write(struct.pack("<hhhhh", *cNode))
    return f.getdata()


class Event:
    def __init__(self, data):
        self.data = Br(io.BytesIO(data))

    def extract(self):
        f = self.data
        f.seek(0xd74)
        pallets = []
        tiles = []
        for i in range(4):
            f.read(12)
            pallets.append(f.read(0x200))
        for i in range(4):
            tileSize = f.readUint16()
            f.read(10)
            tiles.append(f.read(tileSize))
        for i in range(4):
            bpal = Pal(pallets[i])
            palet = bpal.getpal()
            texture = Tile(tiles[i], 64, 240)
            datPng = texture.getpng(palet)
            fileSave(datPng, "HHMBTNI/EVENT_BASE/NPC_SPRITE{0}.png".format(i + 1))

    def insert(self, slotItem, slotNpc,vramPos, sheet):
        # <header>
        newEv = Br(io.BytesIO())
        newEv.write(b"\x80\x88")
        newEv.writeUint16(len(slotItem)+len(slotNpc) + 2)
        filesOffset = []
        files = Br(io.BytesIO())
        for i in range(len(slotItem)):
            newEv.writeUint8(i + 1)
            newEv.writeUint8(8)
        for i in range(len(slotNpc)):
            newEv.writeUint8(i+ 1)
            newEv.writeUint8(32)
        newEv.writeUint16(0x8000)
        newEv.writeUint16(0x4000)
        newEv.wpad(0x34)
        # <init>
        palettes = Br(io.BytesIO())
        tiles = Br(io.BytesIO())
        indexNpc = 0
        totalSpr = 0
        for jsName in slotItem:
            itemName = jsName[13:]
            js = json.loads(open(jsName + "/INFO.json", "r").read())
            for i in range(len(js[itemName]["RAM_INFO"])):
                key = jsName + "/SPRITE/{0:04d}.png".format(i)
                x, y, w, h = js[itemName]["RAM_INFO"][i]
                tile = Br(open(jsName + "/BINARY/{0:04d}.bin".format(i), "r+b"))
                tile.read(2)
                js[itemName]["RAM_INFO"][i] =  (vramPos[key]["x"]//4)+vramPos[key]["ref"], vramPos[key]["y"], w, h
                if h == 1:
                    tile.write(b"\x00\x04")
                    tile.write(struct.pack("<hhhh",  *js[itemName]["RAM_INFO"][i]))
                    palettes.write(tile.getdata())
                    palettes.wpad(4)
                else:
                    tile.read(2)
                    tile.write(struct.pack("<hhhh",  *js[itemName]["RAM_INFO"][i]))
                    tiles.write(tile.getdata())
                    tiles.wpad(4)
                totalSpr += 1
            filesOffset.append(files.tell())
            files.write(Encode(js, itemName))

        for jsName in slotNpc:
            npcName = jsName[12:]
            js = json.loads(open(jsName + "/INFO.json", "r").read())
            for i in range(len(js[npcName]["RAM_INFO"])):
                key = jsName + "/SPRITE/{0:04d}.png".format(i)
                x, y, w, h = js[npcName]["RAM_INFO"][i]
                tile = Br(open(jsName + "/BINARY/{0:04d}.bin".format(i), "r+b"))
                tile.read(2)
                js[npcName]["RAM_INFO"][i] =  (vramPos[key]["x"]//4)+vramPos[key]["ref"], vramPos[key]["y"], w, h
                if h == 1:
                    tile.write(b"\x00\x04")
                    tile.write(struct.pack("<hhhh", *js[npcName]["RAM_INFO"][i]))
                    palettes.write(tile.getdata())
                    palettes.wpad(4)
                else:
                    tile.read(2)
                    tile.write(struct.pack("<hhhh", *js[npcName]["RAM_INFO"][i]))
                    tiles.write(tile.getdata())
                    tiles.wpad(4)
                totalSpr += 1
            filesOffset.append(files.tell())
            files.write(Encode(js, npcName))
        filesOffset.append(files.tell())
        files.write(struct.pack("<HH", 0x100d, totalSpr) + palettes.getdata()+tiles.getdata())
        filesOffset.append(files.tell())
        i = 1
        while True:
            num = sheet.cell(row=i, column=1).value
            s = []
            for j in range(9):
                val = sheet.cell(row=i, column=j + 1).value
                s.append(val)
            newCode = repackCode(s)
            try:
                if (s[1] == "END"):
                    files.write(b"\xff\x00\x00")
                    break
                else:

                    newCode = repackCode(s)
                    files.write(newCode)



            except:
                print(s)
                print("error di line :{0}".format(num))
                sys.exit()

            i += 1

        filesOffset.append(files.tell())
        for i in filesOffset:
            newEv.writeUint32(i)
        newEv.write(files.getdata())

        return newEv.getdata()

def eventInsert(sheet,eventIdx,eventPos,triggerPos,item,npc,vX,vY):
   
    sprites = []
    items =[]
    npcs = []
    i = 1
    while i < npc:
        npcname = sheet.cell(row=1 + i, column=12).value
        if npcname:
            npcs.append("HHMBTNI/NPC/" + npcname)
            dicNpc.update({npcname:i})
            sprites.extend(glob.glob("HHMBTNI/NPC/" + npcname+"/SPRITE/*.png"))
        i += 1
        
    items.append("HHMBTNI/ITEM/0000")
    dicItem.update({"0000":1})
    sprites.extend(glob.glob("HHMBTNI/ITEM/0000/SPRITE/*.png"))
    i = 2
    while i < item:
        itemname = sheet.cell(row=1 + i-1, column=13).value
        if itemname:
            items.append("HHMBTNI/ITEM/" + "{0:04d}".format(DICT_ITEM[itemname]))
            dicItem.update({"{0:04d}".format(DICT_ITEM[itemname]):i-1})
            sprites.extend(glob.glob("HHMBTNI/ITEM/" + "{0:04d}".format(DICT_ITEM[itemname])+"/SPRITE/*.png"))
        i += 1
    vramfile = glob.glob("PyTexturePacker/list/vram*")
    trigpos =0
    for x in vramfile:
        os.remove(x)
        pass
    main.pack_test(sprites)  
    vramjson = glob.glob("PyTexturePacker/list/vram*.json")
    vramPos = {}
    xpos = vX
    for x in vramjson:
         js = json.loads(open(x, "r").read())
         for frame in js["frames"]:
            dictpos = {}
            dictpos.update({"x":js["frames"][frame]["frame"]["x"],"ref":xpos})
            dictpos.update({"y":js["frames"][frame]["frame"]["y"]+vY})
            vramPos.update({frame.replace("\\","/"):dictpos})
         xpos+=64
    data = open("HHMBTNI/EVENT_BASE.bin", "rb").read()
    print("REPACK EVENT...")
    ev = Event(data)
    buffEvent = ev.insert(items, npcs,vramPos, sheet)
    newEvent = io.BytesIO(buffEvent)
    eventSize = len(buffEvent)
    if eventSize > 0x32000:
        print("Max size terlewati!!\nInsert Gagal")
        sys.exit()

    print("REPACK EVENT SIP!")
    iso = open("isos/HHMBTNI_BASE.iso", "r+b")
    iso.seek(eventPos)
    while chunk := newEvent.read(0x800):
        iso.write(chunk)
        iso.read(0x130)
    iso.seek(triggerPos)
    iso.write(struct.pack("H",(eventIdx|0xFC00)))
    iso.write(b"\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\xFF\x00\x00\xB8\xFC\x7F\xFF\xFF\x02\x00")
    iso.write(b"\xFF\xFF")
    
def Main():
    filename = "HHMBTNI/EVENT_BASE/COMMAND.xlsx"
    wb = openpyxl.load_workbook(filename, data_only=True)
    sheets = wb.sheetnames
    i=0
    for nameSheet in sorted(sheets):
        if nameSheet[:5] == "EVENT":
            sheet = wb[nameSheet]
            ins = str(sheet.cell(row=2, column=10).value)
           
            if ins.upper() == "YES":
                idx =int(nameSheet[-2:])
                triggerPos = TRIGGER_POS[sheet.cell(row=2, column=11).value]
                eventIdx = EVENTS_IDX[idx]
                eventPos = EVENTS_POS[idx]
                print("INSERT >>",nameSheet)
                eventInsert(sheet,eventIdx,eventPos,triggerPos,20, 20, -64,256)
                disableDailyNpc = str(sheet.cell(row=4, column=10).value)
                targetSaveState = int(sheet.cell(row=4, column=11).value)
                if idx ==1:
                    
                    if disableDailyNpc == "YES":
                        try:
                            
                            dec = gzip.decompress(open("sstates/SLUS_011.15.{0:03d}".format(targetSaveState-1),"rb").read())
                            dat = io.BytesIO(dec)
                            for dailyPos in disNpc:
                                dat.seek(dailyPos,0)
                                dat.write(b"\x00\x00\x00\x00")
                            dat.seek(0,0)
                            out =  open("sstates/SLUS_011.15.{0:03d}".format(targetSaveState-1),"wb").write(gzip.compress(dat.read()))
                        except:
                            continue
                    else:
                        try:
                            
                            dec = gzip.decompress(open("sstates/SLUS_011.15.{0:03d}".format(targetSaveState-1),"rb").read())
                            dat = io.BytesIO(dec)
                            for dailyPos in disNpc:
                                dat.seek(dailyPos,0)
                                dat.write(b"\xAA\xF5\x02\x0C")
                            dat.seek(0,0)
                            out =  open("sstates/SLUS_011.15.{0:03d}".format(targetSaveState-1),"wb").write(gzip.compress(dat.read()))
                        except:
                            continue
                
            else:
                continue
            
        i+=1
    

if __name__ == "__main__":

   
    try:
        Main()
        insertText()
    except:
        traceback.print_exc()
        input()
        
    
    
  
    #Main(20, 20, -64,256)#NORMAL

